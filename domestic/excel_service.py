"""
Excel (xlsx) generation for the Domestic Tax Invoice — matches the printed
form: company header, invoice meta, Bill-To / Shipping, GST items grid
(SGST/CGST/IGST), totals, amount-in-words, summary box, bank detail, terms.

14-column grid (A..N):
 1 Sl | 2 Description | 3 HS/SAC | 4 Qty | 5 Rate | 6 Amount | 7 Taxable
 8 SGST% | 9 SGST amt | 10 CGST% | 11 CGST amt | 12 IGST% | 13 IGST amt | 14 Total
"""

from io import BytesIO
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NCOLS = 14
# Sl, Desc(trimmed), HS, Qty, Rate, Amount, Taxable, SGST%, SGST amt,
# CGST%, CGST amt, IGST%, IGST amt, Total — GST cols widened.
COL_WIDTHS = [4, 17, 8, 5, 12, 13, 13, 6, 12, 6, 12, 6, 12, 14]
THIN = Side(style='thin', color='000000')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '#,##0.00'


def _m(v):
    return float(Decimal(str(v or 0)))


def _set(ws, r, c, value=None, *, bold=False, size=9.5, align='left', valign='center',
         wrap=True, numfmt=None):
    cell = ws.cell(row=r, column=c)
    cell.value = value
    cell.font = Font(bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if numfmt:
        cell.number_format = numfmt
    return cell


def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _box(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER


def _lv(ws, r, c1, c2, label, value, *, rr=None):
    """Merged 'label : value' cell."""
    _merge(ws, r, c1, rr or r, c2)
    txt = f"{label} {value}" if value not in (None, '') else label
    _set(ws, r, c1, txt, align='left', valign='top')


def _rate(v):
    v = float(v or 0)
    return f"{v:g}%" if v else "-"


def build_tax_invoice_xlsx(ti):
    is_service = ti.kind == 'SERVICE'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tax Invoice'
    for i in range(NCOLS):
        ws.column_dimensions[get_column_letter(i + 1)].width = COL_WIDTHS[i]
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    r = 1
    # ── Company header — copy label top-right, company block centered full width
    _merge(ws, r, 1, r, 14); _set(ws, r, 1, ti.copy_label, size=8, align='right'); r += 1
    _merge(ws, r, 1, r, 14); _set(ws, r, 1, ti.company_name, bold=True, size=15, align='center'); r += 1
    _merge(ws, r, 1, r, 14); _set(ws, r, 1, ti.company_address, size=9.5, align='center'); r += 1
    _merge(ws, r, 1, r, 14); _set(ws, r, 1, f"GSTIN: {ti.company_gstin}    PAN NO. {ti.company_pan}", size=9.5, align='center'); r += 1
    _merge(ws, r, 1, r, 14); _set(ws, r, 1, f"IEC NO. {ti.company_iec}", size=9.5, align='center'); r += 1
    title = 'SERVICE TAX INVOICE' if is_service else 'TAX INVOICE'
    _merge(ws, r, 1, r, 14); _set(ws, r, 1, title, bold=True, size=14, align='center'); _box(ws, r, r, 1, 14); r += 1

    # ── Invoice meta (left A:G, right H:N) ───────────────────────────────────
    meta_top = r
    _lv(ws, r, 1, 7, 'INVOICE NO. :', ti.invoice_no)
    if is_service:
        _lv(ws, r, 8, 14, 'WORK ORDER NO. :', ti.work_order_no)
    else:
        _lv(ws, r, 8, 14, 'VENDOR CODE :', ti.vendor_code)
    r += 1
    _lv(ws, r, 1, 7, 'INVOICE DATE :', ti.invoice_date.strftime('%d.%m.%Y') if ti.invoice_date else '')
    if is_service:
        _lv(ws, r, 8, 14, 'PLACE OF SUPPLY & SERVICE :', ti.place_of_supply)
    else:
        mt = f"VEHICLE NO. : {ti.vehicle_no}    MODE OF TRANSPORT : {ti.mode_of_transport}"
        _lv(ws, r, 8, 14, mt, '')
    r += 1
    if is_service:
        _lv(ws, r, 1, 7, 'WORK ORDER :', ti.work_order_no)
        _lv(ws, r, 8, 14, 'BUYER’S ORDER NO. :', ti.buyers_order_no)
    else:
        ch = f"CHALLAN NO. {ti.challan_no}"
        if ti.challan_date:
            ch += f"  DTD. {ti.challan_date.strftime('%d.%m.%Y')}"
        _lv(ws, r, 1, 7, ch, '')
        _lv(ws, r, 8, 14, 'PLACE OF SUPPLY :', ti.place_of_supply)
    r += 1
    _lv(ws, r, 1, 4, 'STATE :', ti.state)
    _set(ws, r, 5, 'CODE', bold=True, align='center')
    _merge(ws, r, 6, r, 7); _set(ws, r, 6, ti.state_code, align='center')
    bo = ti.buyers_order_no
    if ti.buyers_order_date:
        bo += f"  DTD. {ti.buyers_order_date.strftime('%d.%m.%Y')}"
    _lv(ws, r, 8, 14, 'BUYER’S ORDER NO. :', '' if is_service else bo)
    r += 1
    _box(ws, meta_top, r - 1, 1, 14)

    # ── Bill To / Shipping ───────────────────────────────────────────────────
    bs_top = r
    _merge(ws, r, 1, r, 7); _set(ws, r, 1, 'BILL TO PARTY', bold=True, align='center')
    _merge(ws, r, 8, r, 14); _set(ws, r, 8, 'SHIPPING ADDRESS', bold=True, align='center'); r += 1
    _lv(ws, r, 1, 7, 'NAME :', ti.bill_to_name)
    _lv(ws, r, 8, 14, 'NAME :', ti.ship_to_name); r += 1
    _lv(ws, r, 1, 7, 'ADDRESS :', ti.bill_to_address)
    _lv(ws, r, 8, 14, 'ADDRESS :', ti.ship_to_address); r += 1
    _lv(ws, r, 1, 7, 'GSTIN :', ti.bill_to_gstin)
    _lv(ws, r, 8, 14, 'PROJECT NAME :', ti.ship_to_project); r += 1
    _lv(ws, r, 1, 7, 'STATE/COUNTRY :', ti.bill_to_state)
    _lv(ws, r, 8, 14, 'STATE/COUNTRY :', ti.ship_to_state); r += 1
    _box(ws, bs_top, r - 1, 1, 14)

    # ── Items header (2 rows) ────────────────────────────────────────────────
    it_top = r
    desc_h = 'Work Description' if is_service else 'Product Description'
    code_h = 'SAC Code' if is_service else 'H.S. Code'
    spans = [(1, 'Sl. No.'), (2, desc_h), (3, code_h), (4, 'Qty'), (5, 'Rate'),
             (6, 'Amount'), (7, 'Taxable value'), (14, 'Total Amount (INR)')]
    for c, label in spans:
        _merge(ws, r, c, r + 1, c); _set(ws, r, c, label, bold=True, align='center')
    for c1, c2, label in [(8, 9, 'SGST'), (10, 11, 'CGST'), (12, 13, 'IGST')]:
        _merge(ws, r, c1, r, c2); _set(ws, r, c1, label, bold=True, align='center')
        _set(ws, r + 1, c1, 'Rate', bold=True, align='center')
        _set(ws, r + 1, c2, 'Amount', bold=True, align='center')
    ws.row_dimensions[r].height = 16
    r += 2
    first_item_row = r

    for idx, it in enumerate(ti.items.all(), start=1):
        _set(ws, r, 1, idx, align='center', valign='top')
        _set(ws, r, 2, it.description, align='left', valign='top')
        _set(ws, r, 3, it.hs_sac_code, align='center', valign='top')
        _set(ws, r, 4, f"{_m(it.quantity):g} {it.unit}".strip(), align='center', valign='top')
        _set(ws, r, 5, _m(it.rate), align='right', valign='top', numfmt=MONEY)
        _set(ws, r, 6, _m(it.amount), align='right', valign='top', numfmt=MONEY)
        _set(ws, r, 7, _m(it.taxable_value), align='right', valign='top', numfmt=MONEY)
        _set(ws, r, 8, _rate(it.sgst_rate), align='center', valign='top')
        _set(ws, r, 9, _m(it.sgst_amount), align='right', valign='top', numfmt=MONEY)
        _set(ws, r, 10, _rate(it.cgst_rate), align='center', valign='top')
        _set(ws, r, 11, _m(it.cgst_amount), align='right', valign='top', numfmt=MONEY)
        _set(ws, r, 12, _rate(it.igst_rate), align='center', valign='top')
        if it.igst_amount:
            _set(ws, r, 13, _m(it.igst_amount), align='right', valign='top', numfmt=MONEY)
        else:
            _set(ws, r, 13, '-', align='center', valign='top')
        _set(ws, r, 14, _m(it.total_amount), align='right', valign='top', numfmt=MONEY)
        lines = max(1, str(it.description).count('\n') + 1)
        ws.row_dimensions[r].height = max(18, 14 * lines)
        r += 1

    # filler rows so the grid is tall like the form
    for _ in range(max(3, 10 - (r - first_item_row))):
        ws.row_dimensions[r].height = 16
        r += 1

    # TOTAL row
    _merge(ws, r, 1, r, 5); _set(ws, r, 1, 'TOTAL :', bold=True, align='left')
    _set(ws, r, 6, _m(sum(i.amount for i in ti.items.all())), bold=True, align='right', numfmt=MONEY)
    _set(ws, r, 7, _m(ti.total_amount_before_tax), bold=True, align='right', numfmt=MONEY)
    _set(ws, r, 9, _m(sum(i.sgst_amount for i in ti.items.all())), bold=True, align='right', numfmt=MONEY)
    _set(ws, r, 11, _m(sum(i.cgst_amount for i in ti.items.all())), bold=True, align='right', numfmt=MONEY)
    igst_total = sum(i.igst_amount for i in ti.items.all())
    if igst_total:
        _set(ws, r, 13, _m(igst_total), bold=True, align='right', numfmt=MONEY)
    else:
        _set(ws, r, 13, '-', bold=True, align='center')
    _set(ws, r, 14, _m(ti.total_amount_after_tax), bold=True, align='right', numfmt=MONEY)
    total_row = r
    r += 1
    _box(ws, it_top, total_row, 1, 14)

    # ── Words (left) + summary box (right) ───────────────────────────────────
    sec = r
    _merge(ws, r, 1, r, 7); _set(ws, r, 1, 'TOTAL INVOICE AMOUNT IN WORD', bold=True, align='center')
    _merge(ws, r, 8, r, 12); _set(ws, r, 8, 'TOTAL AMOUNT BEFORE TAX :', bold=True, align='left')
    _set(ws, r, 13, _m(ti.total_amount_before_tax), align='right', numfmt=MONEY); _merge(ws, r, 13, r, 14)
    r += 1
    _merge(ws, r, 1, r + 2, 7); _set(ws, r, 1, ti.amount_in_words, bold=True, align='left', valign='top')
    _merge(ws, r, 8, r, 12); _set(ws, r, 8, 'TOTAL TAX AMOUNT :', bold=True, align='left')
    _set(ws, r, 13, _m(ti.total_tax_amount), align='right', numfmt=MONEY); _merge(ws, r, 13, r, 14)
    r += 1
    _merge(ws, r, 8, r, 12); _set(ws, r, 8, 'TOTAL AMOUNT AFTER TAX (INR) :', bold=True, align='left')
    _set(ws, r, 13, _m(ti.total_amount_after_tax), align='right', numfmt=MONEY); _merge(ws, r, 13, r, 14)
    r += 1
    _merge(ws, r, 8, r, 12); _set(ws, r, 8, 'GST ON REVERSE CHARGE :', bold=True, align='left')
    _merge(ws, r, 13, r, 14); _set(ws, r, 13, ti.gst_on_reverse_charge, align='center')
    r += 1
    _box(ws, sec, r - 1, 1, 14)

    # ── Bank detail (left) + certified/signatory (right) ─────────────────────
    bank = r
    _merge(ws, r, 1, r, 7); _set(ws, r, 1, 'BANK DETAIL', bold=True, align='center')
    _merge(ws, r, 8, r, 14); _set(ws, r, 8, 'Certified that the particulars given above are true and correct', size=9, align='center'); r += 1
    _merge(ws, r, 1, r, 7); _set(ws, r, 1, f"Bank Name : {ti.bank_name}", align='left'); r += 1
    _merge(ws, r, 1, r, 7); _set(ws, r, 1, f"Bank A/C : {ti.bank_account}", align='left')
    _merge(ws, r, 8, r, 14); _set(ws, r, 8, f"For, {ti.company_name}", bold=True, align='center'); r += 1
    _merge(ws, r, 1, r, 7); _set(ws, r, 1, f"Bank IFSC : {ti.bank_ifsc}", align='left'); r += 1
    _merge(ws, r, 1, r, 7); _set(ws, r, 1, 'TERMS OF PAYMENT', bold=True, align='center')
    _merge(ws, r, 8, r, 14); _set(ws, r, 8, 'AUTHORISED SIGNATORY', bold=True, align='center', valign='bottom')
    r += 1
    terms = ti.terms_of_payment or []
    for t in terms:
        _merge(ws, r, 1, r, 7); _set(ws, r, 1, str(t), size=9, align='left'); r += 1
    if not terms:
        _merge(ws, r, 1, r, 7); _set(ws, r, 1, ' '); r += 1
    _box(ws, bank, r - 1, 1, 14)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
