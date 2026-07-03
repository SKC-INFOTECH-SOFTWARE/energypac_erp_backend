"""
Excel export for Proforma Invoices.

Produces a 2-sheet workbook that mirrors the client's template
("PI NO.32 FOR PAPER & SHEET"):

    Sheet "PI"     -> the client-facing Proforma Invoice (export format)
    Sheet "Sheet1" -> the internal Comparative Statement / Note Sheet
                      (vendor purchase cost + price break-up + approvals)

The PI sheet is built entirely from ProformaInvoice data. The Note Sheet
additionally pulls the *purchase* cost from the SELECTED vendor quotation
for each line (PI item -> requisition_item -> selected VendorQuotationItem),
converting the vendor's INR rate into the PI currency via conversion_rate.
Freight / export cost / last-supply comparison come from the new per-line
fields captured on the PI.
"""
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ── Shared styles ───────────────────────────────────────────────────────────
THIN = Side(style='thin', color='000000')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOLD = Font(bold=True, size=9)
BOLD_LG = Font(bold=True, size=12)
NORMAL = Font(size=9)
SMALL = Font(size=8)
HEAD_FILL = PatternFill('solid', fgColor='E5E7EB')
TITLE_FILL = PatternFill('solid', fgColor='F3F4F6')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')


def _d(val):
    """Coerce anything to Decimal safely."""
    try:
        return Decimal(str(val or 0))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0')


def _num(val):
    return float(_d(val))


def _fmt_date(d):
    if not d:
        return ''
    try:
        return d.strftime('%d.%m.%Y')
    except AttributeError:
        return str(d)


# ── Vendor purchase cost lookup ─────────────────────────────────────────────
def _purchase_rate_for_item(pi_item, pi_currency, conversion_rate):
    """
    Return (purchase_unit_price_in_pi_currency, vendor_offers).

    purchase_unit_price -> from the SELECTED vendor quotation for this line,
    converted from the quotation currency into the PI currency.
    vendor_offers -> [{'vendor': name, 'rate': Decimal(pi_currency), 'selected': bool}]
    """
    if not pi_item.requisition_item_id:
        return Decimal('0'), []

    # Imported lazily to avoid app-loading order issues.
    from requisitions.models import VendorQuotationItem

    vq_items = (
        VendorQuotationItem.objects
        .filter(vendor_item__requisition_item_id=pi_item.requisition_item_id)
        .select_related('quotation', 'quotation__assignment__vendor')
    )

    def to_pi_currency(rate, quote_currency):
        rate = _d(rate)
        # Vendor quoted in INR, PI in another currency -> divide by rate.
        if quote_currency == 'INR' and pi_currency != 'INR' and conversion_rate:
            cr = _d(conversion_rate)
            return rate / cr if cr else rate
        return rate

    offers = []
    selected_rate = Decimal('0')
    for vqi in vq_items:
        q = vqi.quotation
        rate = to_pi_currency(vqi.quoted_rate, q.currency)
        vendor_name = ''
        try:
            vendor_name = q.assignment.vendor.vendor_name
        except Exception:
            vendor_name = ''
        offers.append({'vendor': vendor_name, 'rate': rate, 'selected': q.is_selected})
        if q.is_selected:
            selected_rate = rate

    # If nothing explicitly selected, fall back to the cheapest offer.
    if selected_rate == 0 and offers:
        selected_rate = min(o['rate'] for o in offers)

    return selected_rate, offers


def _actual_purchase_rate_for_item(pi_item, pi_currency, conversion_rate):
    """
    Weighted-average ACTUAL purchase unit rate (in PI currency) for this line,
    taken from the real Purchase Order items raised against the same
    requisition + product (non-cancelled). Returns Decimal('0') if never bought.
    """
    if not pi_item.requisition_item_id:
        return Decimal('0')

    from purchase_orders.models import PurchaseOrderItem

    req_id = pi_item.requisition_item.requisition_id
    qs = (
        PurchaseOrderItem.objects
        .filter(po__requisition_id=req_id, product_id=pi_item.product_id)
        .exclude(po__status='CANCELLED')
        .select_related('po')
    )

    tot_qty = Decimal('0')
    tot_amt = Decimal('0')  # in PI currency
    for poi in qs:
        qty = _d(poi.quantity)
        rate = _d(poi.rate)
        po_ccy = poi.po.currency or 'INR'
        # Convert PO rate -> PI currency (same convention as vendor offers).
        if po_ccy == 'INR' and pi_currency != 'INR' and conversion_rate:
            cr = _d(conversion_rate)
            rate = rate / cr if cr else rate
        elif po_ccy != 'INR' and pi_currency == 'INR' and poi.po.conversion_rate:
            rate = rate * _d(poi.po.conversion_rate)
        tot_qty += qty
        tot_amt += rate * qty

    return (tot_amt / tot_qty) if tot_qty > 0 else Decimal('0')


# ── Sheet builders ──────────────────────────────────────────────────────────
def _group_by_lc(computed):
    """
    Reorder note-sheet lines so items sharing the same (non-empty) LC No & Date
    sit next to each other, grouped by first appearance and otherwise stable.
    Items without an LC keep their own spot (never clustered together).
    Lets the user enter matching LC text on any items in any order — the sheet
    groups + merges them automatically.
    """
    order = []
    decorated = []
    for idx, entry in enumerate(computed):
        lc = (entry[0].last_lc_reference or '').strip()
        key = lc if lc else f'__empty_{idx}'  # empty LC -> unique, so it won't group
        if key not in order:
            order.append(key)
        decorated.append((key, idx, entry))
    key_index = {k: i for i, k in enumerate(order)}
    decorated.sort(key=lambda d: (key_index[d[0]], d[1]))
    return [d[2] for d in decorated]


def compute_note_sheet(pi, items):
    """
    Shared data model for the Note Sheet / Comparative Statement (Sheet1).

    Both the Excel export and the landscape PDF render from this so their
    numbers always match. Returns a dict:

        {
          'currency', 'conversion_rate', 'profit_pct',
          'vendor_order': [name, ...],   # dynamic vendor columns
          'generic': bool,               # True when no named vendor quotes
          'rows': [ {sl, description, unit, qty,
                     vendor_rates:[Decimal|None], vendor_totals:[Decimal|None],
                     sale_unit, cpt_total,           # New Sale Price (Kolkata)
                     lc_ref, last_unit, last_total,  # Last Sale Price (Kolkata)
                     p_rate, purchase_total, freight, export_cost, profit, cpt2}, ... ],
          'totals': {vendor:[...], new, last, purchase, freight, export, profit, cpt2},
        }
    """
    conversion_rate = pi.conversion_rate
    pi_currency = pi.currency or 'INR'
    profit_pct = _d(pi.profit_loading_percent)

    computed = []
    for it in items:
        _sel, offers = _purchase_rate_for_item(it, pi_currency, conversion_rate)

        # Apply per-item vendor-offer overrides (edited on the Comparison Sheet).
        overrides = it.vendor_offers if isinstance(it.vendor_offers, dict) else {}
        if overrides:
            existing = {(o.get('vendor') or '').strip(): o for o in offers}
            for vname, vrate in overrides.items():
                key = (vname or '').strip()
                if not key:
                    continue
                orate = _d(vrate)
                if key in existing:
                    existing[key]['rate'] = orate
                    existing[key]['overridden'] = True
                else:
                    offers.append({'vendor': key, 'rate': orate, 'selected': False, 'overridden': True})

        # Recompute the purchase (selected/cheapest) rate from the final offers.
        p_rate = Decimal('0')
        for o in offers:
            if o.get('selected'):
                p_rate = _d(o.get('rate'))
        if p_rate == 0 and offers:
            p_rate = min(_d(o.get('rate')) for o in offers)

        computed.append((it, p_rate, offers))
    computed = _group_by_lc(computed)

    vendor_order = []
    for (_it, _pr, offers) in computed:
        for o in offers:
            name = (o.get('vendor') or '').strip()
            if name and name not in vendor_order:
                vendor_order.append(name)
    generic = not vendor_order
    if generic:
        vendor_order = ['Vendor']
    V = len(vendor_order)

    def vendor_rate(offers, p_rate, name):
        if generic:
            return p_rate if p_rate else None
        for o in offers:
            if (o.get('vendor') or '').strip() == name:
                return _d(o.get('rate'))
        return None

    rows = []
    vend_totals = [Decimal('0')] * V
    tot_new = tot_last = tot_actual = Decimal('0')
    t_purchase = t_freight = t_export = t_profit = t_cpt2 = Decimal('0')

    for idx, (it, p_rate, offers) in enumerate(computed, start=1):
        qty = _d(it.quantity)
        sale_unit = _d(it.unit_price)
        cpt_total = _d(it.amount) or (qty * sale_unit)
        last_up = _d(it.last_unit_price)
        last_total = last_up * qty
        actual_rate = _actual_purchase_rate_for_item(it, pi_currency, conversion_rate)
        actual_total = actual_rate * qty
        tot_new += cpt_total
        tot_last += last_total
        tot_actual += actual_total

        vendor_rates, vendor_totals = [], []
        for i, vname in enumerate(vendor_order):
            vr = vendor_rate(offers, p_rate, vname)
            if vr is None:
                vendor_rates.append(None)
                vendor_totals.append(None)
            else:
                vt = vr * qty
                vend_totals[i] += vt
                vendor_rates.append(vr)
                vendor_totals.append(vt)

        purchase_total = qty * p_rate
        freight = _d(it.freight)
        export_cost = _d(it.export_cost)
        profit = (purchase_total * profit_pct / 100) if profit_pct else Decimal('0')
        cpt2 = purchase_total + freight + export_cost + profit
        t_purchase += purchase_total
        t_freight += freight
        t_export += export_cost
        t_profit += profit
        t_cpt2 += cpt2

        name = it.product.item_name if it.product_id else ''
        prod_desc = (getattr(it.product, 'description', '') or '') if it.product_id else ''
        full_desc = name + (('\n' + prod_desc) if prod_desc and prod_desc != name else '')

        rows.append({
            'sl': idx,
            'description': full_desc,
            'unit': it.unit or (it.product.unit if it.product_id else ''),
            'qty': qty,
            'vendor_rates': vendor_rates,
            'vendor_totals': vendor_totals,
            'actual_rate': actual_rate,
            'actual_total': actual_total,
            'sale_unit': sale_unit,
            'cpt_total': cpt_total,
            'lc_ref': it.last_lc_reference or '',
            'last_unit': last_up,
            'last_total': last_total,
            'p_rate': p_rate,
            'purchase_total': purchase_total,
            'freight': freight,
            'export_cost': export_cost,
            'profit': profit,
            'cpt2': cpt2,
        })

    return {
        'currency': pi_currency,
        'conversion_rate': conversion_rate,
        'profit_pct': profit_pct,
        'vendor_order': vendor_order,
        'generic': generic,
        'rows': rows,
        'totals': {
            'vendor': vend_totals,
            'actual': tot_actual,
            'new': tot_new,
            'last': tot_last,
            'purchase': t_purchase,
            'freight': t_freight,
            'export': t_export,
            'profit': t_profit,
            'cpt2': t_cpt2,
        },
    }


def _merge_consecutive(ws, row_values, column):
    """
    Vertically merge runs of consecutive rows that carry the same non-empty
    value in `column`. `row_values` is a list of (excel_row, value) in order.
    The merged cell keeps the top row's value, centred vertically.
    """
    mid_center = Alignment(horizontal='left', vertical='center', wrap_text=True)
    i = 0
    n = len(row_values)
    while i < n:
        row_i, val = row_values[i]
        j = i
        while val and j + 1 < n and (row_values[j + 1][1] == val):
            j += 1
        if j > i and val:
            ws.merge_cells(start_row=row_i, start_column=column,
                           end_row=row_values[j][0], end_column=column)
            ws.cell(row=row_i, column=column).alignment = mid_center
        i = j + 1


def _style_range(ws, cell_range, font=None, align=None, fill=None, border=True):
    for row in ws[cell_range]:
        for c in row:
            if font:
                c.font = font
            if align:
                c.alignment = align
            if fill:
                c.fill = fill
            if border:
                c.border = BORDER


def _build_pi_sheet(ws, pi, items, symbol):
    widths = [6, 16, 16, 12, 10, 8, 8, 10, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last_col = 'G'

    r = 1
    ws.merge_cells(f'A{r}:{last_col}{r}')
    c = ws[f'A{r}']
    c.value = 'PROFORMA INVOICE'
    c.font = BOLD_LG
    c.alignment = CENTER
    c.fill = TITLE_FILL
    _style_range(ws, f'A{r}:{last_col}{r}')
    ws.row_dimensions[r].height = 22

    # ── Header info grid ────────────────────────────────────────────────
    def label_value(row, col_l, col_r, label, value, merge=True):
        if merge and col_l != col_r:
            ws.merge_cells(f'{col_l}{row}:{col_r}{row}')
        cell = ws[f'{col_l}{row}']
        cell.value = f'{label}\n{value}' if label else value
        cell.font = NORMAL
        cell.alignment = LEFT
        _style_range(ws, f'{col_l}{row}:{col_r}{row}')

    r += 1
    label_value(r, 'A', 'C', 'Exporter / Manufacturer:',
                pi.exporter_beneficiary or 'ENERGYPAC ENGINEERING LIMITED.')
    label_value(r, 'D', 'E', 'Proforma Invoice No. & Date:',
                f'{pi.pi_number}  DT. {_fmt_date(pi.pi_date)}')
    label_value(r, 'F', 'G', 'GST NO.:', pi.gst_number or '')
    ws.row_dimensions[r].height = 46

    r += 1
    label_value(r, 'A', 'C', 'Exporters Ref.:', pi.exporter_reference or '')
    label_value(r, 'D', 'E', 'L/C Number:', pi.lc_number or '')
    label_value(r, 'F', 'G', 'Currency:', pi.currency or '')

    r += 1
    label_value(r, 'A', 'C', 'Consignee:', pi.consignee or '')
    label_value(r, 'D', 'G', 'Applicant / Notify / Importer:', pi.applicant_importer or '')
    ws.row_dimensions[r].height = 40

    r += 1
    label_value(r, 'A', 'B', 'Pre-carriage by:', pi.pre_carriage_by or '')
    label_value(r, 'C', 'C', 'Place of Receipt:', pi.place_of_receipt or '')
    label_value(r, 'D', 'E', 'Country of Origin:', pi.country_of_origin or '')
    label_value(r, 'F', 'G', 'Final Destination:', pi.final_destination or '')
    ws.row_dimensions[r].height = 30

    r += 1
    label_value(r, 'A', 'C', 'Port of Loading:', pi.port_of_loading or '')
    label_value(r, 'D', 'G', 'Port of Discharge:', pi.port_of_discharge or '')

    r += 1
    label_value(r, 'A', 'C', 'Terms of Payment:', pi.terms_of_payment or '')
    label_value(r, 'D', 'G', 'Terms of Delivery:', pi.terms_of_delivery or '')
    ws.row_dimensions[r].height = 26

    # ── Items table ─────────────────────────────────────────────────────
    r += 1
    headers = ['SL #', 'Description of Goods', 'HS CODE NO', 'U.O.M', 'QTY',
               f'UNIT PRICE ({symbol})', f'TOTAL PRICE ({symbol})']
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.font = BOLD
        cell.alignment = CENTER
        cell.fill = HEAD_FILL
        cell.border = BORDER
    ws.row_dimensions[r].height = 26

    subtotal = Decimal('0')
    for idx, it in enumerate(items, start=1):
        r += 1
        desc = it.product.item_name if it.product_id else ''
        qty = _d(it.quantity)
        rate = _d(it.unit_price)
        amt = _d(it.amount) or (qty * rate)
        subtotal += amt
        row_vals = [
            idx, desc, it.hsn_code or '',
            it.unit or (it.product.unit if it.product_id else ''),
            _num(qty), _num(rate), _num(amt),
        ]
        for i, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = BORDER
            cell.font = NORMAL
            if i == 2:
                cell.alignment = LEFT
            elif i in (5, 6, 7):
                cell.alignment = RIGHT
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = CENTER
        ws.row_dimensions[r].height = 30

    # Grand total
    r += 1
    ws.merge_cells(f'A{r}:F{r}')
    tc = ws[f'A{r}']
    tc.value = 'GRAND TOTAL'
    tc.font = BOLD
    tc.alignment = RIGHT
    tc.fill = HEAD_FILL
    grand = _d(pi.grand_total) or subtotal
    gc = ws.cell(row=r, column=7, value=_num(grand))
    gc.font = BOLD
    gc.number_format = '#,##0.00'
    gc.alignment = RIGHT
    gc.fill = HEAD_FILL
    _style_range(ws, f'A{r}:G{r}')

    # ── Terms & Conditions ──────────────────────────────────────────────
    r += 2
    ws.merge_cells(f'A{r}:G{r}')
    tc = ws[f'A{r}']
    tc.value = 'Terms & Conditions:'
    tc.font = BOLD
    tc.alignment = LEFT

    for term in (pi.terms_and_conditions or []):
        r += 1
        if isinstance(term, dict):
            text = f"{term.get('key') or term.get('label', '')}: {term.get('value', '')}"
        else:
            text = str(term)
        ws.merge_cells(f'A{r}:G{r}')
        cell = ws[f'A{r}']
        cell.value = text
        cell.font = SMALL
        cell.alignment = LEFT

    # ── Signature ───────────────────────────────────────────────────────
    r += 3
    ws.merge_cells(f'E{r}:G{r}')
    sc = ws[f'E{r}']
    sc.value = 'For Energypac Engineering Limited'
    sc.font = BOLD
    sc.alignment = CENTER
    r += 3
    ws.merge_cells(f'E{r}:G{r}')
    sc = ws[f'E{r}']
    sc.value = 'Authorized Signatory'
    sc.font = NORMAL
    sc.alignment = CENTER


def _build_note_sheet(ws, pi, items, symbol):
    data = compute_note_sheet(pi, items)
    cur = data['currency']
    vendor_order = data['vendor_order']
    generic = data['generic']
    V = len(vendor_order)
    rows = data['rows']
    tot = data['totals']
    profit_pct = data['profit_pct']

    def numval(v):
        return _num(v) if v is not None else ''

    # ── Column plan (1-based) ───────────────────────────────────────────
    C_SL, C_DESC, C_UOM, C_QTY = 1, 2, 3, 4
    C_VEND0 = 5                              # vendor i -> unit=C_VEND0+2i, total=+1
    C_ACT = C_VEND0 + 2 * V                  # Actual Purchase Price (from PO) — 1 col
    C_NEW = C_ACT + 1                        # New Sale: unit, total
    C_LAST = C_NEW + 2                       # Last Sale: LC, last unit, total
    C_REM = C_LAST + 3                       # Remarks
    LAST_COL = C_REM
    last_letter = get_column_letter(LAST_COL)

    # Column widths
    for col, w in ((C_SL, 6), (C_DESC, 42), (C_UOM, 8), (C_QTY, 8)):
        ws.column_dimensions[get_column_letter(col)].width = w
    for i in range(V):
        ws.column_dimensions[get_column_letter(C_VEND0 + 2 * i)].width = 13
        ws.column_dimensions[get_column_letter(C_VEND0 + 2 * i + 1)].width = 13
    ws.column_dimensions[get_column_letter(C_ACT)].width = 14
    ws.column_dimensions[get_column_letter(C_NEW)].width = 14
    ws.column_dimensions[get_column_letter(C_NEW + 1)].width = 14
    ws.column_dimensions[get_column_letter(C_LAST)].width = 20
    ws.column_dimensions[get_column_letter(C_LAST + 1)].width = 13
    ws.column_dimensions[get_column_letter(C_LAST + 2)].width = 14
    ws.column_dimensions[get_column_letter(C_REM)].width = 16

    def put(row, col, val, font=NORMAL, align=None, fill=None, numfmt=None):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font
        cell.border = BORDER
        if align:
            cell.alignment = align
        if fill:
            cell.fill = fill
        if numfmt:
            cell.number_format = numfmt
        return cell

    # ── Title block ─────────────────────────────────────────────────────
    r = 1
    ws.merge_cells(f'A{r}:{last_letter}{r}')
    c = ws[f'A{r}']
    c.value = 'Comparative Statement / Note Sheet'
    c.font = BOLD_LG
    c.alignment = CENTER
    c.fill = TITLE_FILL

    r += 1
    ws.cell(row=r, column=1, value='PI NO:-').font = BOLD
    ws.merge_cells(f'B{r}:D{r}')
    ws.cell(row=r, column=2, value=f'{pi.pi_number}  DT. {_fmt_date(pi.pi_date)}').font = NORMAL
    r += 1
    ws.cell(row=r, column=1, value='Project Name:').font = BOLD
    ws.merge_cells(f'B{r}:D{r}')
    ws.cell(row=r, column=2, value=pi.project_name or '').font = NORMAL
    r += 1
    ws.cell(row=r, column=1, value='EXCHANGE RATE:').font = BOLD
    ws.cell(row=r, column=3, value=_num(data['conversion_rate'])).font = NORMAL
    ws.cell(row=r, column=4, value=f'PER {cur}').font = NORMAL

    # ── Section 1: grouped 2-row header ─────────────────────────────────
    r += 2
    hr1 = r
    hr2 = r + 1

    # Group-header row
    ws.merge_cells(start_row=hr1, start_column=C_SL, end_row=hr1, end_column=C_QTY)
    put(hr1, C_SL, 'ITEM DETAILS', font=BOLD, align=CENTER, fill=HEAD_FILL)
    for i, vname in enumerate(vendor_order):
        s = C_VEND0 + 2 * i
        ws.merge_cells(start_row=hr1, start_column=s, end_row=hr1, end_column=s + 1)
        title = 'Offer (Ex. works)\nLast Supply' if generic else f'{vname}\nOffer / Last Supply'
        put(hr1, s, title, font=BOLD, align=CENTER, fill=HEAD_FILL)
    ws.merge_cells(start_row=hr1, start_column=C_ACT, end_row=hr2, end_column=C_ACT)
    put(hr1, C_ACT, f'Actual Purchase\nPrice (PO)\n{cur}', font=BOLD, align=CENTER, fill=HEAD_FILL)
    ws.merge_cells(start_row=hr1, start_column=C_NEW, end_row=hr1, end_column=C_NEW + 1)
    put(hr1, C_NEW, 'New Sale Price from Kolkata office', font=BOLD, align=CENTER, fill=HEAD_FILL)
    ws.merge_cells(start_row=hr1, start_column=C_LAST, end_row=hr1, end_column=C_LAST + 2)
    put(hr1, C_LAST, 'Last Sale Price from Kolkata office', font=BOLD, align=CENTER, fill=HEAD_FILL)
    ws.merge_cells(start_row=hr1, start_column=C_REM, end_row=hr2, end_column=C_REM)
    put(hr1, C_REM, 'REMARKS', font=BOLD, align=CENTER, fill=HEAD_FILL)

    # Sub-header row
    put(hr2, C_SL, 'Sl. No.', font=BOLD, align=CENTER, fill=HEAD_FILL)
    put(hr2, C_DESC, 'Description', font=BOLD, align=CENTER, fill=HEAD_FILL)
    put(hr2, C_UOM, 'U.O.M', font=BOLD, align=CENTER, fill=HEAD_FILL)
    put(hr2, C_QTY, 'QTY', font=BOLD, align=CENTER, fill=HEAD_FILL)
    for i in range(V):
        s = C_VEND0 + 2 * i
        put(hr2, s, f'Unit Price\n(Ex. works)\n{cur}', font=BOLD, align=CENTER, fill=HEAD_FILL)
        put(hr2, s + 1, f'Total Price\n(Ex. works)\n{cur}', font=BOLD, align=CENTER, fill=HEAD_FILL)
    put(hr2, C_NEW, f'Unit Price\n(CPT, BENAPOLE)\n{cur}', font=BOLD, align=CENTER, fill=HEAD_FILL)
    put(hr2, C_NEW + 1, f'Total Price\n(CPT, BENAPOLE)\n{cur}', font=BOLD, align=CENTER, fill=HEAD_FILL)
    put(hr2, C_LAST, 'LC NO & DATE', font=BOLD, align=CENTER, fill=HEAD_FILL)
    put(hr2, C_LAST + 1, 'LAST UNIT PRICE', font=BOLD, align=CENTER, fill=HEAD_FILL)
    put(hr2, C_LAST + 2, f'Total Price\n(CPT, BENAPOLE)\n{cur}', font=BOLD, align=CENTER, fill=HEAD_FILL)

    # Border every header cell (incl. merged spans) for a clean grid.
    for rr in (hr1, hr2):
        for cc in range(1, LAST_COL + 1):
            ws.cell(row=rr, column=cc).border = BORDER
    ws.row_dimensions[hr1].height = 30
    ws.row_dimensions[hr2].height = 44

    # ── Section 1: data rows ────────────────────────────────────────────
    r = hr2
    lc_rows = []  # (excel_row, lc_reference) -> merge shared LC cells
    for row in rows:
        r += 1
        lc_rows.append((r, (row['lc_ref'] or '').strip()))
        put(r, C_SL, row['sl'], align=CENTER)
        put(r, C_DESC, row['description'], align=LEFT)
        put(r, C_UOM, row['unit'], align=CENTER)
        put(r, C_QTY, _num(row['qty']), align=RIGHT, numfmt='#,##0.00')
        for i in range(V):
            s = C_VEND0 + 2 * i
            put(r, s, numval(row['vendor_rates'][i]), align=RIGHT, numfmt='#,##0.00')
            put(r, s + 1, numval(row['vendor_totals'][i]), align=RIGHT, numfmt='#,##0.00')
        put(r, C_ACT, (_num(row['actual_rate']) if row['actual_rate'] else ''), align=RIGHT, numfmt='#,##0.00')
        put(r, C_NEW, _num(row['sale_unit']), align=RIGHT, numfmt='#,##0.00')
        put(r, C_NEW + 1, _num(row['cpt_total']), align=RIGHT, numfmt='#,##0.00')
        put(r, C_LAST, row['lc_ref'] or '', align=LEFT)
        put(r, C_LAST + 1, _num(row['last_unit']), align=RIGHT, numfmt='#,##0.00')
        put(r, C_LAST + 2, _num(row['last_total']), align=RIGHT, numfmt='#,##0.00')
        put(r, C_REM, '', align=LEFT)
        ws.row_dimensions[r].height = 40

    # Merge consecutive rows sharing the same (non-empty) LC No & Date.
    _merge_consecutive(ws, lc_rows, column=C_LAST)

    # Totals row
    r += 1
    ws.merge_cells(start_row=r, start_column=C_SL, end_row=r, end_column=C_QTY)
    put(r, C_SL, 'Total Amount', font=BOLD, align=RIGHT, fill=HEAD_FILL)
    for cc in range(C_SL, C_QTY + 1):
        ws.cell(row=r, column=cc).fill = HEAD_FILL
        ws.cell(row=r, column=cc).border = BORDER
    for i in range(V):
        s = C_VEND0 + 2 * i
        put(r, s, '', fill=HEAD_FILL)
        put(r, s + 1, _num(tot['vendor'][i]), font=BOLD, align=RIGHT, fill=HEAD_FILL, numfmt='#,##0.00')
    put(r, C_ACT, _num(tot['actual']), font=BOLD, align=RIGHT, fill=HEAD_FILL, numfmt='#,##0.00')
    put(r, C_NEW, '', fill=HEAD_FILL)
    put(r, C_NEW + 1, _num(tot['new']), font=BOLD, align=RIGHT, fill=HEAD_FILL, numfmt='#,##0.00')
    put(r, C_LAST, '', fill=HEAD_FILL)
    put(r, C_LAST + 1, '', fill=HEAD_FILL)
    put(r, C_LAST + 2, _num(tot['last']), font=BOLD, align=RIGHT, fill=HEAD_FILL, numfmt='#,##0.00')
    put(r, C_REM, '', fill=HEAD_FILL)

    # ── Section 2: Price break-up (same rows, override-aware) ───────────
    r += 3
    ws.merge_cells(f'A{r}:D{r}')
    c = ws[f'A{r}']
    c.value = 'PROFORMA INVOICE NO:-  price breakup based on selected offer:'
    c.font = BOLD
    ws.merge_cells(f'E{r}:F{r}')
    ws.cell(row=r, column=5, value=f'{pi.pi_number}  DT. {_fmt_date(pi.pi_date)}').font = NORMAL

    r += 1
    headers2 = [
        'Sl. No.', 'Description', 'U.O.M', 'QTY',
        f'Purchase Unit Price\n(Ex. works)\n{cur}', f'Total Purchase\n(Ex. works)\n{cur}',
        'Freight', 'Export Cost', f'Profit Loading\n@ {_num(profit_pct):g}%',
        f'Total Amount\n(CPT)\n{cur}',
    ]
    for i, h in enumerate(headers2, start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.font = BOLD
        cell.alignment = CENTER
        cell.fill = HEAD_FILL
        cell.border = BORDER
    ws.row_dimensions[r].height = 42

    for row in rows:
        r += 1
        vals = [
            row['sl'], row['description'], row['unit'],
            _num(row['qty']), _num(row['p_rate']), _num(row['purchase_total']),
            _num(row['freight']), _num(row['export_cost']), _num(row['profit']), _num(row['cpt2']),
        ]
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = BORDER
            cell.font = NORMAL
            if i == 2:
                cell.alignment = LEFT
            elif i in (4, 5, 6, 7, 8, 9, 10):
                cell.alignment = RIGHT
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = CENTER
        ws.row_dimensions[r].height = 34

    r += 1
    ws.merge_cells(f'A{r}:D{r}')
    ws.cell(row=r, column=1, value='Total Amount').font = BOLD
    ws[f'A{r}'].alignment = RIGHT
    for col, val in ((6, tot['purchase']), (7, tot['freight']), (8, tot['export']), (9, tot['profit']), (10, tot['cpt2'])):
        cell = ws.cell(row=r, column=col, value=_num(val))
        cell.number_format = '#,##0.00'
        cell.font = BOLD
        cell.alignment = RIGHT
    _style_range(ws, f'A{r}:J{r}')

    # ── Approvals ───────────────────────────────────────────────────────
    r += 3
    ws.cell(row=r, column=1, value='Negotiated By:').font = BOLD
    ws.cell(row=r, column=3, value='Checked By:').font = BOLD
    r += 2
    ws.cell(row=r, column=1, value='_______________________').font = NORMAL
    ws.cell(row=r, column=3, value='_______________________').font = NORMAL
    r += 1
    ws.cell(row=r, column=1, value=pi.negotiated_by or '').font = BOLD
    ws.cell(row=r, column=3, value=pi.checked_by or '').font = BOLD


def build_pi_workbook(pi):
    """Return an in-memory .xlsx (BytesIO) for the given ProformaInvoice."""
    items = list(pi.items.select_related('product', 'requisition_item').all())
    currency = pi.currency or 'INR'
    symbol = '₹' if currency == 'INR' else ('$' if currency == 'USD' else currency)

    wb = Workbook()
    pi_ws = wb.active
    pi_ws.title = 'PI'
    _build_pi_sheet(pi_ws, pi, items, symbol)

    note_ws = wb.create_sheet('Sheet1')
    _build_note_sheet(note_ws, pi, items, symbol)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
