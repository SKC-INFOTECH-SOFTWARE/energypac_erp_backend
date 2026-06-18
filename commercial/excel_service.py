"""
Excel (xlsx) generation for Commercial Invoice & Packing List.

Mirrors the printed/PDF format: a bordered grid with the same header block for
both documents. In the items table there are NO horizontal lines between items
(only continuous vertical column separators), and each item's HS CODE line is
rendered bold + centered under its description.

A 12-column grid (A..L) is used.
"""

from io import BytesIO
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter


# ── grid / style constants ───────────────────────────────────────────────────
NCOLS = 12
COL_WIDTHS = [7, 9, 8, 8, 10, 10, 8, 9, 7, 9, 10, 14]   # A..L

THIN = Side(style='thin', color='000000')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Item columns: Marks A:B | Pkgs C:D | Desc E:H | Qty I | Unit/wt J:K | Total L
# Vertical separators sit on the right edge of these columns:
RIGHT_SEP = {2, 4, 8, 9, 11, 12}

LABEL_FONT = InlineFont(b=True, sz=8)
VALUE_FONT = InlineFont(sz=8)


def _money(v):
    return float(Decimal(str(v or 0)))


def label_value(label, value):
    blocks = [TextBlock(LABEL_FONT, label)]
    if value not in (None, ''):
        blocks.append(TextBlock(VALUE_FONT, f"\n{value}"))
    return CellRichText(*blocks)


def _set(ws, row, col, value=None, *, bold=False, size=8, align='left',
         valign='top', wrap=True, rich=None):
    cell = ws.cell(row=row, column=col)
    cell.value = rich if rich is not None else value
    cell.font = Font(bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    return cell


def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _border_region(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER


def _vline_row(ws, r, *, top=False, bottom=False):
    """Vertical column separators for one row; horizontal lines only if asked."""
    for c in range(1, NCOLS + 1):
        ws.cell(row=r, column=c).border = Border(
            left=THIN if c == 1 else None,
            right=THIN if c in RIGHT_SEP else None,
            top=THIN if top else None,
            bottom=THIN if bottom else None,
        )


# ── title / header ────────────────────────────────────────────────────────────

def _write_title(ws, title, row):
    _merge(ws, row, 1, row, NCOLS)
    _set(ws, row, 1, title, bold=True, size=13, align='center', valign='center', wrap=False)
    ws.row_dimensions[row].height = 22


def _write_header(ws, ci, start_row):
    r = start_row
    g = lambda v: v or ''   # noqa: E731

    inv_no_date = ci.invoice_no or ''
    if ci.invoice_date:
        inv_no_date = f"{inv_no_date}  Dt. {ci.invoice_date.strftime('%d.%m.%Y')}".strip()
    bo = ci.buyers_order_no or ''
    if ci.buyers_order_date:
        bo = f"{bo}  Dt. {ci.buyers_order_date.strftime('%d.%m.%Y')}".strip()

    _merge(ws, r, 1, r + 2, 6)
    _set(ws, r, 1, rich=label_value('Exporter:', g(ci.exporter)))
    _merge(ws, r, 7, r, 8); _set(ws, r, 7, rich=label_value('Invoice No & Date', inv_no_date))
    _merge(ws, r, 9, r, 10); _set(ws, r, 9, rich=label_value('Exporters Ref.', g(ci.exporters_ref)))
    _merge(ws, r, 11, r, 12); _set(ws, r, 11, rich=label_value('GST NO.', g(ci.gst_no)))
    _merge(ws, r + 1, 7, r + 2, 12)
    _set(ws, r + 1, 7, rich=label_value('Buyers Order No. & Date:', bo))
    r += 3

    _merge(ws, r, 1, r + 2, 6)
    _set(ws, r, 1, rich=label_value('Consigned to the order of:', g(ci.consigned_to_order_of)))
    _merge(ws, r, 7, r, 12)
    _set(ws, r, 7, rich=label_value('Terms of Delivery:', g(ci.terms_of_delivery)))
    _merge(ws, r + 1, 7, r + 2, 12)
    _set(ws, r + 1, 7, rich=label_value('Applicant:', g(ci.applicant)))
    r += 3

    _merge(ws, r, 1, r + 2, 6)
    _set(ws, r, 1, rich=label_value('Importer/Notify Party:', g(ci.importer_notify_party)))
    _merge(ws, r, 7, r + 3, 12)
    _set(ws, r, 7, rich=label_value('Terms of Delivery and Payment:', g(ci.terms_of_delivery_and_payment)))
    r += 3

    _merge(ws, r, 1, r, 6)
    _set(ws, r, 1, rich=label_value('Place of Supply:', g(ci.place_of_supply)))
    r += 1

    cols = [(1, 3), (4, 6), (7, 9), (10, 12)]
    quad = [
        ('Vessel/Flight No:', ci.vessel_flight_no),
        ('Port of Loading:', ci.port_of_loading),
        ('Port of Discharge:', ci.port_of_discharge),
        ('Place of Delivery:', ci.place_of_delivery),
    ]
    for (lbl, val), (c1, c2) in zip(quad, cols):
        _merge(ws, r, c1, r, c2); _set(ws, r, c1, rich=label_value(lbl, g(val)))
    r += 1

    quad2 = [
        ('Pre-carriage by:', ci.pre_carriage_by),
        ('Place of Receipt of by Pre-carriage:', ci.place_of_receipt),
        ('Country of Origin:', ci.country_of_origin),
        ('Final Destination:', ci.final_destination),
    ]
    for (lbl, val), (c1, c2) in zip(quad2, cols):
        _merge(ws, r, c1, r, c2); _set(ws, r, c1, rich=label_value(lbl, g(val)))
    r += 1

    _border_region(ws, start_row, r - 1, 1, NCOLS)
    return r


# ── items table ───────────────────────────────────────────────────────────────

def _items_header(ws, r, col5, col6):
    _merge(ws, r, 1, r, 2); _set(ws, r, 1, 'Marks & Nos/\nContainer No.', bold=True, align='center', valign='center')
    _merge(ws, r, 3, r, 4); _set(ws, r, 3, 'No & Kind of\nPkgs.', bold=True, align='center', valign='center')
    _merge(ws, r, 5, r, 8); _set(ws, r, 5, 'Description of Goods', bold=True, align='center', valign='center')
    _set(ws, r, 9, 'Qty\nSet./Nos.', bold=True, align='center', valign='center')
    _merge(ws, r, 10, r, 11); _set(ws, r, 10, col5, bold=True, align='center', valign='center')
    _set(ws, r, 12, col6, bold=True, align='center', valign='center')
    ws.row_dimensions[r].height = 34
    _vline_row(ws, r, top=True, bottom=True)
    return r + 1


def _item_block(ws, r, pkgs, desc, hs, value_cells):
    """
    Renders columns C..L for one item (Marks A:B is a single merged range for the
    whole document, applied separately). HS code is bold + centered under desc.
    value_cells: list of (text, c1, c2, numfmt, align).
    """
    has_hs = bool(hs)
    r2 = r + 1 if has_hs else r

    _merge(ws, r, 3, r2, 4); _set(ws, r, 3, pkgs, align='left', valign='top')
    _merge(ws, r, 5, r, 8); _set(ws, r, 5, desc, align='left', valign='top')
    if has_hs:
        _merge(ws, r2, 5, r2, 8)
        _set(ws, r2, 5, f"HS CODE NO. {hs}", bold=True, align='center', valign='top')

    for text, c1, c2, numfmt, align in value_cells:
        if c1 != c2 or r2 != r:
            _merge(ws, r, c1, r2, c2)
        cell = _set(ws, r, c1, text, align=align, valign='center')
        if numfmt:
            cell.number_format = numfmt

    for rr in range(r, r2 + 1):
        _vline_row(ws, rr)

    lines = max(1, str(desc).count('\n') + 1)
    ws.row_dimensions[r].height = max(16, 13 * lines)
    if has_hs:
        ws.row_dimensions[r2].height = 15
    return r2 + 1


def _marks_text(ci):
    f, t = (ci.marks_from or '').strip(), (ci.marks_to or '').strip()
    if f and t:
        return f"{f}\nTO\n{t}"
    return f or t or ''


def _setup_sheet(wb, title):
    ws = wb.active
    ws.title = title
    for i in range(NCOLS):
        ws.column_dimensions[get_column_letter(i + 1)].width = COL_WIDTHS[i]
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    return ws


def _finalize(wb):
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _hs_of(item):
    return getattr(item, 'hs_code', '') or ''


# ─────────────────────────────────────────────────────────────────────────────
# Commercial Invoice
# ─────────────────────────────────────────────────────────────────────────────

def build_commercial_invoice_xlsx(ci):
    wb = Workbook()
    ws = _setup_sheet(wb, 'Commercial Invoice')

    _write_title(ws, 'COMMERCIAL INVOICE', 1)
    r = _write_header(ws, ci, 2)

    cur = ci.currency or 'US$'
    r = _items_header(ws, r, f'Unit\nPrice\n{cur}', f'Total\nAmount\n{cur}')
    first_item_row = r

    for it in ci.items.all():
        qty_txt = f"{_money(it.quantity):g} {it.unit}".strip()
        r = _item_block(ws, r, it.no_kind_pkgs, it.description, _hs_of(it), [
            (qty_txt, 9, 9, None, 'center'),
            (_money(it.unit_price), 10, 11, '#,##0.00', 'right'),
            (_money(it.total_amount), 12, 12, '#,##0.00', 'right'),
        ])

    # Filler rows so the table is tall and column lines run continuously
    used = r - first_item_row
    for _ in range(max(2, 18 - used)):
        _vline_row(ws, r); ws.row_dimensions[r].height = 14; r += 1

    # Totals — bottom of the table, values in the description column area
    for lbl, val in [
        ('Total FCA Value (USD):', _money(ci.total_fca_value)),
        ('Total Freight (USD):', _money(ci.total_freight)),
        ('Total CPT Value (USD):', _money(ci.total_cpt_value)),
    ]:
        _merge(ws, r, 3, r, 7); _set(ws, r, 3, lbl, bold=True, align='right')
        cell = _set(ws, r, 8, val, align='right'); cell.number_format = '#,##0.00'
        _vline_row(ws, r)
        r += 1

    # Marks & Nos — one tall merged range cell for the whole shipment
    _merge(ws, first_item_row, 1, r - 1, 2)
    _set(ws, first_item_row, 1, _marks_text(ci), align='center', valign='center')

    # Chargeable amount in words — bordered band closing the table
    _merge(ws, r, 1, r, 11)
    _set(ws, r, 1, rich=label_value('Total Chargeable amount (US$):', ' ' + (ci.amount_in_words or '')))
    cell = _set(ws, r, 12, _money(ci.total_cpt_value), bold=True, align='right')
    cell.number_format = '#,##0.00'
    _vline_row(ws, r, top=True, bottom=True)
    r += 1

    # Footer: declarations | signature
    footer_top = r
    decl = ci.declarations or []
    _set(ws, r, 1, rich=label_value('PROJECT NAME :', ' ' + (ci.project_name or '')))
    _merge(ws, r, 1, r, 8)
    _merge(ws, r, 9, footer_top + len(decl) + 1, 12)
    _set(ws, r, 9, 'Signature', align='center', valign='top')
    r += 1
    for idx, d in enumerate(decl, start=1):
        _merge(ws, r, 1, r, 8); _set(ws, r, 1, f"({idx}) {d}", align='left'); r += 1
    if ci.lut_no:
        _merge(ws, r, 1, r, 8); _set(ws, r, 1, f"LUT NO. {ci.lut_no}", bold=True, align='left'); r += 1
    _border_region(ws, footer_top, max(r, footer_top + len(decl) + 1) - 1, 1, NCOLS)

    return _finalize(wb)


# ─────────────────────────────────────────────────────────────────────────────
# Packing List
# ─────────────────────────────────────────────────────────────────────────────

def build_packing_list_xlsx(pl):
    ci = pl.commercial_invoice
    wb = Workbook()
    ws = _setup_sheet(wb, 'Packing List')

    _write_title(ws, 'PACKING LIST', 1)
    r = _write_header(ws, ci, 2)

    r = _items_header(ws, r, 'NETT\nWEIGHT\nKGS.', 'GROSS\nWEIGHT\nKGS.')
    first_item_row = r

    for it in pl.items.all():
        qty_txt = f"{_money(it.quantity):g} {it.unit}".strip()
        r = _item_block(ws, r, it.no_kind_pkgs, it.description, _hs_of(it), [
            (qty_txt, 9, 9, None, 'center'),
            (float(it.nett_weight or 0), 10, 11, '#,##0.000" Kgs."', 'right'),
            (float(it.gross_weight or 0), 12, 12, '#,##0.000" Kgs."', 'right'),
        ])

    # Filler rows so column lines run full height
    used = r - first_item_row
    for _ in range(max(2, 18 - used)):
        _vline_row(ws, r); ws.row_dimensions[r].height = 14; r += 1

    # Totals row — top + bottom border closing the table (no label, per the form)
    _merge(ws, r, 3, r, 9); _set(ws, r, 3, '')
    c = _set(ws, r, 10, float(pl.total_nett_weight or 0), bold=True, align='right')
    c.number_format = '#,##0.000" Kgs."'; _merge(ws, r, 10, r, 11)
    c = _set(ws, r, 12, float(pl.total_gross_weight or 0), bold=True, align='right')
    c.number_format = '#,##0.000" Kgs."'
    _vline_row(ws, r, top=True, bottom=True)

    # Marks & Nos — one tall merged range cell
    _merge(ws, first_item_row, 1, r, 2)
    _set(ws, first_item_row, 1, _marks_text(ci), align='center', valign='center')
    r += 1

    # Footer
    footer_top = r
    _merge(ws, r, 9, r + 2, 12); _set(ws, r, 9, 'Signature', align='center', valign='top')
    _merge(ws, r, 1, r, 8); _set(ws, r, 1, 'The goods are of Indian Origin', bold=True, align='left'); r += 1
    _merge(ws, r, 1, r, 8); _set(ws, r, 1, rich=label_value('Packing Specification:', ' ' + (pl.packing_specification or ''))); r += 1
    if pl.lut_no:
        _merge(ws, r, 1, r, 8); _set(ws, r, 1, f"LUT NO. {pl.lut_no}", bold=True, align='left'); r += 1
    _border_region(ws, footer_top, max(r, footer_top + 3) - 1, 1, NCOLS)

    return _finalize(wb)
